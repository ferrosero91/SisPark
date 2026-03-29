# Python standard library
from datetime import datetime, timedelta

# Django core
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.timezone import now

# Django database
from django.db.models import Avg, Count, F, Sum
from django.db.models.functions import TruncDate

# Django views
from django.views.generic import CreateView, ListView, TemplateView
from django.views.generic.edit import DeleteView

# Local imports
from .forms import CategoryForm, ParkingTicketForm
from .models import ParkingTicket, VehicleCategory, Caja

# Tenant imports
from tenants.context import get_current_tenant, set_current_tenant
from third_parties.models import ThirdPartyVehicle
from monthly_contracts.models import MonthlyContract


def get_tenant_from_user(user):
    if hasattr(user, 'tenant') and user.tenant:
        return user.tenant
    return None


@login_required
def pagina_inicial(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'registration/login.html')


class CategoryListView(ListView):
    model = VehicleCategory
    template_name = 'parking/category_list.html'
    
    def get_queryset(self):
        tenant = get_tenant_from_user(self.request.user)
        if tenant:
            return VehicleCategory.objects.all_tenants().filter(tenant=tenant)
        return VehicleCategory.objects.none()


class CategoryCreateView(CreateView):
    model = VehicleCategory
    form_class = CategoryForm
    template_name = 'parking/category_form.html'
    success_url = reverse_lazy('category-list')

    def form_valid(self, form):
        tenant = get_tenant_from_user(self.request.user)
        if tenant:
            set_current_tenant(tenant)
            form.instance.tenant = tenant
            messages.success(self.request, "Categoría creada con éxito.")
            return super().form_valid(form)
        else:
            messages.error(self.request, "No tiene un parqueadero asignado.")
            return self.form_invalid(form)

    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                if field == '__all__':
                    messages.error(self.request, error)
                else:
                    field_name = form.fields.get(field).label if field in form.fields else field
                    messages.error(self.request, f"{field_name}: {error}")
        return super().form_invalid(form)


class VehicleEntryView(CreateView):
    model = ParkingTicket
    form_class = ParkingTicketForm
    template_name = 'parking/vehicle_entry.html'
    success_url = reverse_lazy('print-ticket')

    def dispatch(self, request, *args, **kwargs):
        # Verificar si tiene turno activo
        tenant = get_tenant_from_user(request.user)
        if tenant and not has_active_turno(request.user, tenant):
            messages.warning(request, 'Debe abrir un turno antes de registrar vehículos.')
            return redirect('abrir_turno')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['tenant'] = get_tenant_from_user(self.request.user)
        return kwargs

    def form_valid(self, form):
        try:
            tenant = get_tenant_from_user(self.request.user)
            if tenant:
                set_current_tenant(tenant)
                form.instance.tenant = tenant
            
            category = form.cleaned_data['category']
            placa = form.cleaned_data['placa'].upper()
            
            if tenant:
                # Usar el servicio de detección de contratos
                from monthly_contracts.services import ContractDetectionService, RateType
                
                detection_service = ContractDetectionService(tenant)
                rate_result = detection_service.detect_rate(placa, category)
                
                # Asignar tercero y contrato si existe
                if rate_result.third_party:
                    form.instance.third_party = rate_result.third_party
                
                if rate_result.vehicle:
                    # Guardar referencia al vehículo registrado
                    pass
                
                if rate_result.contract:
                    form.instance.monthly_contract = rate_result.contract
                    form.instance.is_monthly_entry = True
                
                # Guardar info de tarifa especial en sesión para mostrar en ticket
                if rate_result.rate_type != RateType.REGULAR:
                    self.request.session['rate_info'] = {
                        'type': rate_result.rate_type.value,
                        'message': rate_result.message,
                        'days_remaining': rate_result.days_remaining
                    }
            
            if category.name.upper() in ['MOTOS', 'MOTO']:
                cascos = self.request.POST.get('cascos', 0)
                form.instance.cascos = int(cascos) if cascos else 0
            
            response = super().form_valid(form)
            self.request.session['ticket_id'] = str(self.object.id)
            return response
        except IntegrityError:
            messages.error(self.request, 'Este vehículo ya se encuentra en el estacionamiento.')
            return self.form_invalid(form)


@login_required
def vehicle_exit(request):
    tenant = get_tenant_from_user(request.user)
    
    # Verificar si tiene turno activo
    if tenant and not has_active_turno(request.user, tenant):
        messages.warning(request, 'Debe abrir un turno antes de registrar salidas.')
        return redirect('abrir_turno')
    
    if request.method == 'POST':
        identifier = request.POST.get('identifier')
        action = request.POST.get('action', 'search')
        
        ticket_query = ParkingTicket.objects.all_tenants().filter(
            placa__iexact=identifier,
            exit_time=None
        )
        
        if tenant:
            ticket_query = ticket_query.filter(tenant=tenant)
        
        ticket = ticket_query.first()

        if ticket:
            # Solo buscar - no procesar salida aún
            if action == 'search':
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    # Usar servicio de cálculo de tarifas
                    from monthly_contracts.services import RateCalculationService
                    
                    rate_service = RateCalculationService(tenant)
                    rate_info = rate_service.calculate_exit_fee(ticket)
                    
                    duration = ticket.get_current_duration()
                    duration_str = f"{duration['hours']}h {duration['minutes']}m"
                    
                    return JsonResponse({
                        'amount': float(rate_info['amount']),
                        'original_amount': float(rate_info.get('original_amount', rate_info['amount'])),
                        'discount': float(rate_info.get('discount', 0)),
                        'duration': duration_str,
                        'placa': ticket.placa,
                        'entry_time': ticket.entry_time.strftime('%d/%m/%Y %H:%M'),
                        'ticket_id': str(ticket.id),
                        'is_monthly': ticket.is_monthly_entry or rate_info['is_contract'],
                        'rate_type': rate_info['rate_type'],
                        'rate_message': rate_info['message'],
                        'days_remaining': rate_info.get('days_remaining', 0),
                        'category': ticket.category.name if ticket.category else '',
                        'marca': ticket.marca or '',
                        'color': ticket.color or '',
                        'third_party': ticket.third_party.full_name if ticket.third_party else None,
                    })
            
            # Confirmar salida - ahora sí procesar
            elif action == 'confirm':
                try:
                    ticket.exit_time = timezone.now()
                    ticket.amount_paid = ticket.calculate_fee()
                    ticket.save()
                    
                    return JsonResponse({
                        'success': True,
                        'ticket_id': str(ticket.id),
                        'amount': float(ticket.amount_paid)
                    })
                except Exception as e:
                    return JsonResponse({'error': str(e)}, status=500)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': 'Vehículo no encontrado'}, status=404)
        messages.error(request, 'Vehículo no encontrado')
        return redirect('vehicle-exit')

    placa = request.GET.get('placa', '')
    
    return render(request, 'parking/vehicle_exit.html', {
        'placa': placa
    })


@login_required
def vehicle_payment(request):
    """Vista para procesar el pago de un vehículo"""
    tenant = get_tenant_from_user(request.user)
    ticket_id = request.GET.get('ticket_id')
    
    if not ticket_id:
        messages.error(request, 'No se especificó un ticket')
        return redirect('vehicle-exit')
    
    # Limpiar el ticket_id de separadores de miles y convertir a entero
    try:
        if isinstance(ticket_id, str):
            ticket_id = ticket_id.replace('.', '').replace(',', '')
        ticket_id = int(ticket_id)
    except (ValueError, TypeError):
        messages.error(request, 'ID de ticket inválido')
        return redirect('vehicle-exit')
    
    ticket_query = ParkingTicket.objects.all_tenants().filter(id=ticket_id, exit_time=None)
    if tenant:
        ticket_query = ticket_query.filter(tenant=tenant)
    ticket = ticket_query.first()
    
    if not ticket:
        messages.error(request, 'Ticket no encontrado o ya procesado')
        return redirect('vehicle-exit')
    
    # Calcular duración y monto
    duration = ticket.get_current_duration()
    duration_str = f"{duration['hours']}h {duration['minutes']}m"
    amount = ticket.calculate_fee()
    
    # Obtener métodos de pago
    from parking.models_config import PaymentMethod
    payment_methods = []
    first_is_cash = False
    if tenant:
        payment_methods = list(PaymentMethod.objects.all_tenants().filter(
            tenant=tenant, 
            is_active=True, 
            allow_for_exit=True,
            is_credit=False
        ).order_by('order', 'name'))
        if payment_methods:
            first_is_cash = payment_methods[0].payment_type == 'cash'
    
    return render(request, 'parking/vehicle_payment.html', {
        'ticket': ticket,
        'duration': duration_str,
        'amount': amount,
        'payment_methods': payment_methods,
        'first_is_cash': first_is_cash
    })


@login_required
def print_exit_ticket(request):
    tenant = get_tenant_from_user(request.user)
    
    if request.method == 'POST':
        ticket_id = request.POST.get('ticket_id')
        amount_received = request.POST.get('amount_received')
        payment_method_id = request.POST.get('payment_method')
        is_cash = request.POST.get('is_cash') == '1'
        
        # Limpiar ticket_id de cualquier formato de número
        if ticket_id:
            ticket_id = str(ticket_id).replace('.', '').replace(',', '').strip()

        if ticket_id:
            try:
                from parking.models_config import PaymentMethod
                
                ticket_query = ParkingTicket.objects.all_tenants().filter(id=ticket_id)
                if tenant:
                    ticket_query = ticket_query.filter(tenant=tenant)
                ticket = ticket_query.first()
                
                if not ticket:
                    messages.error(request, 'Ticket no encontrado')
                    return redirect('dashboard')
                
                # Obtener método de pago
                payment_method = None
                if payment_method_id:
                    payment_method = PaymentMethod.objects.all_tenants().filter(
                        id=payment_method_id, tenant=tenant
                    ).first()
                
                # Si el ticket no tiene exit_time, procesar la salida ahora
                if not ticket.exit_time:
                    ticket.exit_time = timezone.now()
                    ticket.amount_paid = ticket.calculate_fee()
                    ticket.payment_method = payment_method
                    ticket.save()
                
                # Validar y convertir amount_received de forma robusta
                try:
                    if amount_received is None or str(amount_received).strip() == '':
                        amount_received = float(ticket.amount_paid or 0)
                    else:
                        amount_received = float(amount_received)
                except (ValueError, TypeError, AttributeError):
                    amount_received = float(ticket.amount_paid or 0)
                
                amount_paid = float(ticket.amount_paid or 0)
                change = amount_received - amount_paid if is_cash else 0

                return render(request, 'parking/print_exit_ticket.html', {
                    'ticket': ticket,
                    'parking_lot': tenant,
                    'amount_received': amount_received,
                    'change': change,
                    'is_cash': is_cash,
                    'current_time': timezone.now(),
                })
            except Exception as e:
                messages.error(request, f'Error al imprimir ticket: {str(e)}')
                return redirect('dashboard')

    messages.warning(request, 'No se especificó un ticket para imprimir')
    return redirect('dashboard')


@login_required
def dashboard(request):
    tenant = get_tenant_from_user(request.user)
    today = timezone.now().date()
    start_date = today - timedelta(days=7)
    
    if tenant:
        base_query = ParkingTicket.objects.all_tenants().filter(tenant=tenant)
    else:
        base_query = ParkingTicket.objects.none()
    
    recent_tickets = base_query.filter(
        entry_time__date__gte=start_date,
        entry_time__date__lte=today
    )

    recent_revenue = base_query.filter(
        exit_time__date__gte=start_date,
        exit_time__date__lte=today,
        exit_time__isnull=False,
        amount_paid__isnull=False
    )

    active_vehicles = base_query.filter(
        exit_time__isnull=True
    ).select_related('category')

    daily_stats = base_query.filter(
        exit_time__date__gte=start_date,
        exit_time__date__lte=today,
        exit_time__isnull=False,
        amount_paid__isnull=False
    ).values('exit_time__date').annotate(
        revenue=Sum('amount_paid'),
        count=Count('id')
    ).order_by('exit_time__date')

    category_stats = base_query.filter(
        exit_time__date__gte=start_date,
        exit_time__date__lte=today,
        exit_time__isnull=False,
        amount_paid__isnull=False
    ).values('category__name').annotate(
        count=Count('id'),
        revenue=Sum('amount_paid')
    )

    # Usar servicio de contratos para obtener resumen
    contracts_expiring = []
    contract_summary = {}
    if tenant:
        from monthly_contracts.services import ContractDetectionService, get_contract_summary
        
        detection_service = ContractDetectionService(tenant)
        contracts_expiring = detection_service.check_expiring_contracts(days=5)
        contract_summary = get_contract_summary(tenant)

    stats = {
        'daily': {
            'total_vehicles': recent_tickets.count(),
            'total_revenue': recent_revenue.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
        },
        'category': category_stats,
        'active_vehicles': active_vehicles.count(),
        'active_vehicles_list': active_vehicles
    }

    context = {
        'stats': stats,
        'daily_stats': [
            {
                'date': stat['exit_time__date'].strftime('%d/%m/%Y'),
                'revenue': stat['revenue'] or 0,
                'count': stat['count'] or 0,
            }
            for stat in daily_stats
        ],
        'current_time': timezone.now(),
        'tenant': tenant,
        'contracts_expiring': contracts_expiring,
        'contract_summary': contract_summary
    }

    return render(request, 'parking/dashboard.html', context)


@login_required
def print_ticket(request):
    tenant = get_tenant_from_user(request.user)
    ticket_id = request.GET.get('ticket_id') or request.session.get('ticket_id')
    
    if ticket_id:
        try:
            # Limpiar el ticket_id de separadores de miles y convertir a entero
            if isinstance(ticket_id, str):
                ticket_id = ticket_id.replace('.', '').replace(',', '')
            ticket_id = int(ticket_id)
            
            ticket_query = ParkingTicket.objects.all_tenants().filter(id=ticket_id)
            if tenant:
                ticket_query = ticket_query.filter(tenant=tenant)
            ticket = ticket_query.first()
            
            if not ticket:
                messages.error(request, 'Ticket no encontrado')
                return redirect('dashboard')
            
            if 'ticket_id' in request.session:
                del request.session['ticket_id']
            
            return render(request, 'parking/print_ticket.html', {
                'ticket': ticket,
                'parking_lot': tenant,
                'is_reprint': bool(request.GET.get('ticket_id')),
                'current_time': timezone.now(),
                'duration': ticket.get_duration() if ticket.exit_time else ticket.get_current_duration(),
                'current_fee': ticket.amount_paid if ticket.exit_time else ticket.calculate_current_fee()
            })
        except Exception as e:
            messages.error(request, f'Error al imprimir ticket: {str(e)}')
            return redirect('dashboard')
    
    messages.warning(request, 'No se especificó un ticket para imprimir')
    return redirect('vehicle-entry')


class ReportView(TemplateView):
    template_name = 'parking/reports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = get_tenant_from_user(self.request.user)
        report_type = self.request.GET.get('type', 'general')
        period = self.request.GET.get('period', 'today')
        page = self.request.GET.get('page', 1)

        # Usar hora local de Colombia
        today = timezone.localtime(timezone.now())
        
        # Manejar períodos predefinidos
        if period == 'yesterday':
            yesterday = today - timedelta(days=1)
            start_date = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif period == 'week':
            start_of_week = today - timedelta(days=today.weekday())
            start_date = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif period == 'month':
            start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif period == 'year':
            start_date = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif period == 'custom':
            start_str = self.request.GET.get('start_date')
            end_str = self.request.GET.get('end_date')
            if start_str and end_str:
                start_date = timezone.make_aware(datetime.strptime(f"{start_str} 00:00:00", '%Y-%m-%d %H:%M:%S'))
                end_date = timezone.make_aware(datetime.strptime(f"{end_str} 23:59:59", '%Y-%m-%d %H:%M:%S'))
            else:
                start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
                end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        else:  # today (default)
            period = 'today'
            start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)

        if tenant:
            tickets_qs = ParkingTicket.objects.all_tenants().filter(
                tenant=tenant,
                exit_time__isnull=False,
                exit_time__range=(start_date, end_date)
            ).exclude(amount_paid__isnull=True).select_related('category', 'payment_method')
        else:
            tickets_qs = ParkingTicket.objects.none()

        # Resumen general (usando agregación, no carga todos los registros)
        summary = tickets_qs.aggregate(
            total_vehicles=Count('id'),
            total_revenue=Sum('amount_paid'),
            avg_duration=Avg(F('exit_time') - F('entry_time')),
            avg_revenue=Avg('amount_paid')
        )

        if summary['avg_duration'] is not None:
            summary['avg_duration'] = summary['avg_duration'].total_seconds() / 3600

        # Por categoría (agregación)
        category_stats = list(tickets_qs.values('category__name').annotate(
            count=Count('id'),
            revenue=Sum('amount_paid'),
        ).order_by('-revenue'))

        # Por método de pago (agregación optimizada)
        from collections import defaultdict
        from decimal import Decimal
        from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
        
        payment_method_stats = list(tickets_qs.values(
            'payment_method__name', 
            'payment_method__icon',
            'payment_method__payment_type'
        ).annotate(
            count=Count('id'),
            total=Sum('amount_paid')
        ).order_by('-total'))
        
        payment_stats_list = []
        for stat in payment_method_stats:
            payment_stats_list.append({
                'name': stat['payment_method__name'] or 'Efectivo',
                'icon': stat['payment_method__icon'] or 'fa-money-bill',
                'is_cash': stat['payment_method__payment_type'] == 'cash' if stat['payment_method__payment_type'] else True,
                'count': stat['count'],
                'total': stat['total'] or Decimal('0')
            })

        # Ingresos diarios (agregación)
        daily_stats = list(tickets_qs.annotate(
            date=TruncDate('exit_time')
        ).values('date').annotate(
            count=Count('id'),
            revenue=Sum('amount_paid')
        ).order_by('-date'))

        # Vehículos frecuentes (limitado a 10)
        frequent_vehicles = list(tickets_qs.values('placa').annotate(
            visits=Count('id'),
            total_spent=Sum('amount_paid')
        ).order_by('-visits')[:10])

        # Mensualidades
        from monthly_contracts.models import MonthlyContract, ContractPayment
        
        contract_payments = ContractPayment.objects.filter(
            tenant=tenant,
            payment_date__range=(start_date, end_date),
            is_confirmed=True
        ).select_related('payment_method', 'contract__third_party').prefetch_related('contract__vehicles__vehicle') if tenant else []
        
        contracts_revenue = sum(p.amount for p in contract_payments)
        
        # Cuentas por cobrar
        today_date = now().date()
        pending_contracts = MonthlyContract.objects.all_tenants().filter(
            tenant=tenant,
            is_active=True,
            end_date__lt=today_date
        ).select_related('third_party').prefetch_related('vehicles__vehicle') if tenant else []
        
        pending_total = sum(c.monthly_rate for c in pending_contracts)
        
        # Contratos activos
        active_contracts_count = MonthlyContract.objects.all_tenants().filter(
            tenant=tenant,
            is_active=True,
            status='active'
        ).count() if tenant else 0

        # Historial de salidas con PAGINACIÓN
        tickets_ordered = tickets_qs.order_by('-exit_time')
        paginator = Paginator(tickets_ordered, 10)  # 10 por página
        
        try:
            recent_exits = paginator.page(page)
        except PageNotAnInteger:
            recent_exits = paginator.page(1)
        except EmptyPage:
            recent_exits = paginator.page(paginator.num_pages)

        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'period': period,
            'report_type': report_type,
            'summary': summary,
            'category_stats': category_stats,
            'payment_stats': payment_stats_list,
            'daily_stats': daily_stats,
            'frequent_vehicles': frequent_vehicles,
            'contract_payments': contract_payments,
            'contracts_revenue': contracts_revenue,
            'pending_contracts': pending_contracts,
            'pending_total': pending_total,
            'active_contracts_count': active_contracts_count,
            'recent_exits': recent_exits,
            'parking_lot': tenant,
            'total_general': float(summary['total_revenue'] or 0) + float(contracts_revenue),
        })
        
        # Gastos del período
        from .models import Expense
        if tenant:
            expenses = Expense.objects.all_tenants().filter(
                tenant=tenant,
                date__gte=start_date.date(),
                date__lte=end_date.date()
            ).select_related('category', 'created_by').order_by('-date', '-created_at')
            
            total_expenses = sum(float(e.amount) for e in expenses)
            
            # Agrupar gastos por categoría
            expenses_by_category = defaultdict(lambda: {'count': 0, 'total': Decimal('0')})
            for e in expenses:
                expenses_by_category[e.category.name]['count'] += 1
                expenses_by_category[e.category.name]['total'] += e.amount
        else:
            expenses = []
            total_expenses = 0
            expenses_by_category = {}
        
        context.update({
            'expenses': expenses,
            'total_expenses': total_expenses,
            'expenses_by_category': dict(expenses_by_category),
            'net_total': float(summary['total_revenue'] or 0) + float(contracts_revenue) - total_expenses,
        })
        
        return context


@login_required
def export_report_excel(request):
    """Exportar reporte completo a Excel con gastos y utilidad"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from django.http import HttpResponse
    from collections import defaultdict
    from decimal import Decimal
    
    tenant = get_tenant_from_user(request.user)
    start_str = request.GET.get('start_date', timezone.now().date().isoformat())
    end_str = request.GET.get('end_date', timezone.now().date().isoformat())
    
    try:
        start_date = timezone.make_aware(datetime.strptime(f"{start_str} 00:00:00", '%Y-%m-%d %H:%M:%S'))
        end_date = timezone.make_aware(datetime.strptime(f"{end_str} 23:59:59", '%Y-%m-%d %H:%M:%S'))
    except ValueError:
        start_date = end_date = timezone.now()
    
    # Obtener datos de tickets
    tickets = ParkingTicket.objects.all_tenants().filter(
        tenant=tenant,
        exit_time__isnull=False,
        exit_time__range=(start_date, end_date)
    ).exclude(amount_paid__isnull=True).select_related('category', 'payment_method').order_by('-exit_time')
    
    from monthly_contracts.models import ContractPayment, MonthlyContract
    contract_payments = ContractPayment.objects.filter(
        tenant=tenant,
        payment_date__range=(start_date, end_date),
        is_confirmed=True
    ).select_related('payment_method', 'contract__third_party').prefetch_related('contract__vehicles__vehicle')
    
    # Contratos vencidos
    pending_contracts = MonthlyContract.objects.all_tenants().filter(
        tenant=tenant,
        is_active=True,
        end_date__lt=timezone.now().date()
    ).select_related('third_party').prefetch_related('vehicles__vehicle')
    
    # Obtener gastos
    from .models import Expense
    expenses = Expense.objects.all_tenants().filter(
        tenant=tenant,
        date__gte=start_date.date(),
        date__lte=end_date.date()
    ).select_related('category', 'created_by').order_by('-date', '-created_at')
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=14, color="1E293B")
    subtitle_font = Font(bold=True, size=11, color="64748B")
    section_font = Font(bold=True, size=12, color="1E293B")
    total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    total_font = Font(bold=True, size=11)
    currency_green = Font(bold=True, color="059669")
    currency_red = Font(bold=True, color="DC2626")
    expense_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
    profit_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Calcular totales
    total_tickets = sum(float(t.amount_paid or 0) for t in tickets)
    total_contracts = sum(float(p.amount or 0) for p in contract_payments)
    total_pending = sum(float(c.monthly_rate or 0) for c in pending_contracts)
    total_expenses = sum(float(e.amount or 0) for e in expenses)
    total_income = total_tickets + total_contracts
    net_profit = total_income - total_expenses
    
    # ========== HOJA 1: RESUMEN EJECUTIVO ==========
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    
    # Título
    ws1.merge_cells('A1:D1')
    ws1['A1'] = f"REPORTE FINANCIERO - {tenant.name if tenant else 'SoluPark'}"
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    ws1.merge_cells('A2:D2')
    ws1['A2'] = f"Período: {start_str} al {end_str}"
    ws1['A2'].font = subtitle_font
    ws1['A2'].alignment = Alignment(horizontal='center')
    
    ws1.merge_cells('A3:D3')
    ws1['A3'] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws1['A3'].font = Font(size=9, color="94A3B8")
    ws1['A3'].alignment = Alignment(horizontal='center')
    
    # ========== SECCIÓN: INGRESOS ==========
    row = 5
    ws1[f'A{row}'] = "INGRESOS"
    ws1[f'A{row}'].font = section_font
    
    row += 1
    headers = ['Concepto', 'Cantidad', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    income_data = [
        ('Tickets de Parqueadero', tickets.count(), total_tickets),
        ('Pagos de Mensualidades', contract_payments.count(), total_contracts),
    ]
    
    for item in income_data:
        row += 1
        ws1.cell(row=row, column=1, value=item[0]).border = border
        ws1.cell(row=row, column=2, value=item[1]).border = border
        cell = ws1.cell(row=row, column=3, value=item[2])
        cell.number_format = '"$"#,##0'
        cell.border = border
    
    row += 1
    ws1.cell(row=row, column=1, value='TOTAL INGRESOS').border = border
    ws1.cell(row=row, column=2, value=tickets.count() + contract_payments.count()).border = border
    cell = ws1.cell(row=row, column=3, value=total_income)
    cell.number_format = '"$"#,##0'
    cell.border = border
    for c in range(1, 4):
        ws1.cell(row=row, column=c).fill = total_fill
        ws1.cell(row=row, column=c).font = total_font
    
    # ========== SECCIÓN: GASTOS ==========
    row += 2
    ws1[f'A{row}'] = "GASTOS"
    ws1[f'A{row}'].font = section_font
    
    # Agrupar gastos por categoría
    expenses_by_cat = defaultdict(lambda: {'count': 0, 'total': Decimal('0')})
    for e in expenses:
        expenses_by_cat[e.category.name]['count'] += 1
        expenses_by_cat[e.category.name]['total'] += e.amount or Decimal('0')
    
    row += 1
    headers = ['Categoría', 'Cantidad', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.border = border
    
    for cat_name, cat_data in sorted(expenses_by_cat.items(), key=lambda x: -float(x[1]['total'])):
        row += 1
        ws1.cell(row=row, column=1, value=cat_name).border = border
        ws1.cell(row=row, column=2, value=cat_data['count']).border = border
        cell = ws1.cell(row=row, column=3, value=float(cat_data['total']))
        cell.number_format = '"$"#,##0'
        cell.border = border
        cell.font = currency_red
    
    row += 1
    ws1.cell(row=row, column=1, value='TOTAL GASTOS').border = border
    ws1.cell(row=row, column=2, value=expenses.count()).border = border
    cell = ws1.cell(row=row, column=3, value=total_expenses)
    cell.number_format = '"$"#,##0'
    cell.border = border
    for c in range(1, 4):
        ws1.cell(row=row, column=c).fill = expense_fill
        ws1.cell(row=row, column=c).font = Font(bold=True, color="DC2626")
    
    # ========== SECCIÓN: UTILIDAD NETA ==========
    row += 2
    ws1.merge_cells(f'A{row}:B{row}')
    ws1[f'A{row}'] = "UTILIDAD NETA (Ingresos - Gastos)"
    ws1[f'A{row}'].font = Font(bold=True, size=12)
    cell = ws1.cell(row=row, column=3, value=net_profit)
    cell.number_format = '"$"#,##0'
    cell.font = currency_green if net_profit >= 0 else currency_red
    for c in range(1, 4):
        ws1.cell(row=row, column=c).fill = profit_fill if net_profit >= 0 else expense_fill
        ws1.cell(row=row, column=c).border = border
    
    # ========== SECCIÓN: MÉTODOS DE PAGO ==========
    row += 3
    ws1[f'A{row}'] = "INGRESOS POR MÉTODO DE PAGO"
    ws1[f'A{row}'].font = section_font
    
    payment_summary = defaultdict(lambda: {'count': 0, 'total': Decimal('0')})
    for t in tickets:
        method = t.payment_method.name if t.payment_method else 'Efectivo'
        payment_summary[method]['count'] += 1
        payment_summary[method]['total'] += t.amount_paid or Decimal('0')
    for p in contract_payments:
        method = p.payment_method.name if p.payment_method else 'Sin especificar'
        payment_summary[method]['count'] += 1
        payment_summary[method]['total'] += p.amount or Decimal('0')
    
    row += 1
    headers = ['Método de Pago', 'Operaciones', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
        cell.border = border
    
    for method, pdata in sorted(payment_summary.items(), key=lambda x: -float(x[1]['total'])):
        row += 1
        ws1.cell(row=row, column=1, value=method).border = border
        ws1.cell(row=row, column=2, value=pdata['count']).border = border
        cell = ws1.cell(row=row, column=3, value=float(pdata['total']))
        cell.number_format = '"$"#,##0'
        cell.border = border
    
    # Ajustar anchos
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 18
    
    # ========== HOJA 2: PARQUEADERO ==========
    ws2 = wb.create_sheet("Parqueadero")
    
    ws2.merge_cells('A1:H1')
    ws2['A1'] = "DETALLE DE TICKETS DE PARQUEADERO"
    ws2['A1'].font = title_font
    
    headers = ['Fecha', 'Hora Salida', 'Placa', 'Categoría', 'Entrada', 'Duración', 'Método Pago', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    row = 3
    for ticket in tickets:
        row += 1
        duration = (ticket.exit_time - ticket.entry_time).total_seconds() / 3600
        ws2.cell(row=row, column=1, value=ticket.exit_time.strftime('%d/%m/%Y')).border = border
        ws2.cell(row=row, column=2, value=ticket.exit_time.strftime('%H:%M')).border = border
        ws2.cell(row=row, column=3, value=ticket.placa).border = border
        ws2.cell(row=row, column=4, value=ticket.category.name if ticket.category else '-').border = border
        ws2.cell(row=row, column=5, value=ticket.entry_time.strftime('%H:%M')).border = border
        ws2.cell(row=row, column=6, value=f"{duration:.1f}h").border = border
        ws2.cell(row=row, column=7, value=ticket.payment_method.name if ticket.payment_method else 'Efectivo').border = border
        cell = ws2.cell(row=row, column=8, value=float(ticket.amount_paid or 0))
        cell.number_format = '"$"#,##0'
        cell.border = border
    
    # Total
    row += 1
    ws2.cell(row=row, column=7, value="TOTAL").font = total_font
    cell = ws2.cell(row=row, column=8, value=total_tickets)
    cell.number_format = '"$"#,##0'
    cell.font = currency_green
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws2.column_dimensions[col].width = 14
    
    # ========== HOJA 3: MENSUALIDADES ==========
    ws3 = wb.create_sheet("Mensualidades")
    
    ws3.merge_cells('A1:F1')
    ws3['A1'] = "PAGOS DE MENSUALIDADES"
    ws3['A1'].font = title_font
    
    headers = ['Fecha', 'Cliente', 'Vehículos/Combo', 'Período', 'Método Pago', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    row = 3
    for payment in contract_payments:
        row += 1
        ws3.cell(row=row, column=1, value=payment.payment_date.strftime('%d/%m/%Y %H:%M')).border = border
        ws3.cell(row=row, column=2, value=str(payment.contract.third_party) if payment.contract else '-').border = border
        
        # Vehículos o combo
        if payment.contract and payment.contract.use_combo_rate and payment.contract.combo_name:
            vehicles_text = f"[COMBO] {payment.contract.combo_name}"
        else:
            vehicles_text = payment.contract.vehicles_list if payment.contract else '-'
        ws3.cell(row=row, column=3, value=vehicles_text).border = border
        
        ws3.cell(row=row, column=4, value=f"{payment.get_month_name()} {payment.payment_year}").border = border
        ws3.cell(row=row, column=5, value=payment.payment_method.name if payment.payment_method else '-').border = border
        cell = ws3.cell(row=row, column=6, value=float(payment.amount or 0))
        cell.number_format = '"$"#,##0'
        cell.border = border
    
    # Total
    row += 1
    ws3.cell(row=row, column=5, value="TOTAL").font = total_font
    cell = ws3.cell(row=row, column=6, value=total_contracts)
    cell.number_format = '"$"#,##0'
    cell.font = currency_green
    
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 25
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 15
    ws3.column_dimensions['F'].width = 15
    
    # ========== HOJA 4: GASTOS ==========
    ws4 = wb.create_sheet("Gastos")
    
    ws4.merge_cells('A1:F1')
    ws4['A1'] = "DETALLE DE GASTOS"
    ws4['A1'].font = title_font
    
    headers = ['Fecha', 'Categoría', 'Descripción', 'Registrado por', 'Método Pago', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.border = border
    
    row = 3
    for expense in expenses:
        row += 1
        ws4.cell(row=row, column=1, value=expense.date.strftime('%d/%m/%Y')).border = border
        ws4.cell(row=row, column=2, value=expense.category.name).border = border
        ws4.cell(row=row, column=3, value=expense.description or '-').border = border
        ws4.cell(row=row, column=4, value=expense.created_by.get_full_name() if expense.created_by else '-').border = border
        ws4.cell(row=row, column=5, value=expense.payment_method).border = border
        cell = ws4.cell(row=row, column=6, value=float(expense.amount or 0))
        cell.number_format = '"$"#,##0'
        cell.border = border
    
    # Total
    row += 1
    ws4.cell(row=row, column=5, value="TOTAL GASTOS").font = total_font
    cell = ws4.cell(row=row, column=6, value=total_expenses)
    cell.number_format = '"$"#,##0'
    cell.font = currency_red
    
    ws4.column_dimensions['A'].width = 12
    ws4.column_dimensions['B'].width = 18
    ws4.column_dimensions['C'].width = 35
    ws4.column_dimensions['D'].width = 18
    ws4.column_dimensions['E'].width = 12
    ws4.column_dimensions['F'].width = 15
    
    # ========== HOJA 5: POR COBRAR ==========
    ws5 = wb.create_sheet("Por Cobrar")
    
    ws5.merge_cells('A1:E1')
    ws5['A1'] = "CONTRATOS VENCIDOS - CUENTAS POR COBRAR"
    ws5['A1'].font = title_font
    
    headers = ['Cliente', 'Teléfono', 'Vehículos/Combo', 'Vencimiento', 'Tarifa Mensual']
    for col, header in enumerate(headers, 1):
        cell = ws5.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        cell.border = border
    
    row = 3
    for contract in pending_contracts:
        row += 1
        ws5.cell(row=row, column=1, value=contract.third_party.full_name).border = border
        ws5.cell(row=row, column=2, value=contract.third_party.phone or '-').border = border
        
        if contract.use_combo_rate and contract.combo_name:
            vehicles_text = f"[COMBO] {contract.combo_name}"
        else:
            vehicles_text = contract.vehicles_list or 'Sin vehículos'
        ws5.cell(row=row, column=3, value=vehicles_text).border = border
        
        ws5.cell(row=row, column=4, value=contract.end_date.strftime('%d/%m/%Y')).border = border
        cell = ws5.cell(row=row, column=5, value=float(contract.monthly_rate or 0))
        cell.number_format = '"$"#,##0'
        cell.border = border
    
    # Total
    row += 1
    ws5.cell(row=row, column=4, value="TOTAL POR COBRAR").font = total_font
    cell = ws5.cell(row=row, column=5, value=total_pending)
    cell.number_format = '"$"#,##0'
    cell.font = Font(bold=True, color="F59E0B")
    
    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 15
    ws5.column_dimensions['C'].width = 30
    ws5.column_dimensions['D'].width = 15
    ws5.column_dimensions['E'].width = 18
    
    # Respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=reporte_{start_str}_a_{end_str}.xlsx'
    wb.save(response)
    return response


@login_required
def export_report_pdf(request):
    """Exportar reporte completo a PDF con gastos y utilidad"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from django.http import HttpResponse
    from io import BytesIO
    from collections import defaultdict
    from decimal import Decimal
    
    tenant = get_tenant_from_user(request.user)
    start_str = request.GET.get('start_date', timezone.now().date().isoformat())
    end_str = request.GET.get('end_date', timezone.now().date().isoformat())
    
    try:
        start_date = timezone.make_aware(datetime.strptime(f"{start_str} 00:00:00", '%Y-%m-%d %H:%M:%S'))
        end_date = timezone.make_aware(datetime.strptime(f"{end_str} 23:59:59", '%Y-%m-%d %H:%M:%S'))
    except ValueError:
        start_date = end_date = timezone.now()
    
    # Obtener datos
    tickets = ParkingTicket.objects.all_tenants().filter(
        tenant=tenant,
        exit_time__isnull=False,
        exit_time__range=(start_date, end_date)
    ).exclude(amount_paid__isnull=True).select_related('category', 'payment_method')
    
    total_tickets = sum(float(t.amount_paid or 0) for t in tickets)
    
    from monthly_contracts.models import ContractPayment, MonthlyContract
    contract_payments = ContractPayment.objects.filter(
        tenant=tenant,
        payment_date__range=(start_date, end_date),
        is_confirmed=True
    ).select_related('payment_method', 'contract__third_party')
    total_contracts = sum(float(p.amount or 0) for p in contract_payments)
    
    pending_contracts = MonthlyContract.objects.all_tenants().filter(
        tenant=tenant,
        is_active=True,
        end_date__lt=timezone.now().date()
    ).select_related('third_party').prefetch_related('vehicles__vehicle')
    total_pending = sum(float(c.monthly_rate or 0) for c in pending_contracts)
    
    # Obtener gastos
    from .models import Expense
    expenses = Expense.objects.all_tenants().filter(
        tenant=tenant,
        date__gte=start_date.date(),
        date__lte=end_date.date()
    ).select_related('category', 'created_by').order_by('-date', '-created_at')
    total_expenses = sum(float(e.amount or 0) for e in expenses)
    
    # Calcular totales
    total_income = total_tickets + total_contracts
    net_profit = total_income - total_expenses
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=5, textColor=colors.HexColor('#1E293B'))
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1, spaceAfter=20, textColor=colors.HexColor('#64748B'))
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, spaceBefore=20, spaceAfter=10, textColor=colors.HexColor('#1E293B'))
    
    # Colores
    header_bg = colors.HexColor('#1E293B')
    header_text = colors.white
    total_bg = colors.HexColor('#F1F5F9')
    border_color = colors.HexColor('#E2E8F0')
    green_text = colors.HexColor('#059669')
    red_text = colors.HexColor('#DC2626')
    expense_bg = colors.HexColor('#FEF2F2')
    profit_bg = colors.HexColor('#ECFDF5')
    
    # ========== ENCABEZADO ==========
    elements.append(Paragraph(f"REPORTE FINANCIERO", title_style))
    elements.append(Paragraph(f"{tenant.name if tenant else 'SoluPark'}", subtitle_style))
    elements.append(Paragraph(f"Período: {start_str} al {end_str}", subtitle_style))
    
    # ========== RESUMEN DE INGRESOS ==========
    elements.append(Paragraph("INGRESOS", section_style))
    
    income_data = [
        ['Concepto', 'Cantidad', 'Total'],
        ['Tickets de Parqueadero', str(tickets.count()), f"${total_tickets:,.0f}"],
        ['Pagos de Mensualidades', str(contract_payments.count()), f"${total_contracts:,.0f}"],
        ['TOTAL INGRESOS', str(tickets.count() + contract_payments.count()), f"${total_income:,.0f}"],
    ]
    
    income_table = Table(income_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    income_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), header_text),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), total_bg),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, border_color),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(income_table)
    
    # ========== RESUMEN DE GASTOS ==========
    elements.append(Paragraph("GASTOS", section_style))
    
    # Agrupar gastos por categoría
    expenses_by_cat = defaultdict(lambda: {'count': 0, 'total': Decimal('0')})
    for e in expenses:
        expenses_by_cat[e.category.name]['count'] += 1
        expenses_by_cat[e.category.name]['total'] += e.amount or Decimal('0')
    
    expense_data = [['Categoría', 'Cantidad', 'Total']]
    for cat_name, cat_data in sorted(expenses_by_cat.items(), key=lambda x: -float(x[1]['total'])):
        expense_data.append([cat_name, str(cat_data['count']), f"${float(cat_data['total']):,.0f}"])
    expense_data.append(['TOTAL GASTOS', str(expenses.count()), f"${total_expenses:,.0f}"])
    
    expense_table = Table(expense_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
    expense_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
        ('TEXTCOLOR', (0, 0), (-1, 0), header_text),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), expense_bg),
        ('TEXTCOLOR', (-1, -1), (-1, -1), red_text),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, border_color),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(expense_table)
    
    # ========== UTILIDAD NETA ==========
    elements.append(Spacer(1, 15))
    profit_data = [['UTILIDAD NETA (Ingresos - Gastos)', f"${net_profit:,.0f}"]]
    profit_table = Table(profit_data, colWidths=[4.5*inch, 1.5*inch])
    profit_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), profit_bg if net_profit >= 0 else expense_bg),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (-1, -1), (-1, -1), green_text if net_profit >= 0 else red_text),
        ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 1, border_color),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(profit_table)
    
    # ========== INGRESOS POR MÉTODO DE PAGO ==========
    elements.append(Paragraph("INGRESOS POR MÉTODO DE PAGO", section_style))
    
    payment_summary = defaultdict(lambda: {'count': 0, 'total': Decimal('0')})
    for t in tickets:
        method = t.payment_method.name if t.payment_method else 'Efectivo'
        payment_summary[method]['count'] += 1
        payment_summary[method]['total'] += t.amount_paid or Decimal('0')
    for p in contract_payments:
        method = p.payment_method.name if p.payment_method else 'Sin especificar'
        payment_summary[method]['count'] += 1
        payment_summary[method]['total'] += p.amount or Decimal('0')
    
    payment_data = [['Método de Pago', 'Operaciones', 'Total']]
    for method, data in sorted(payment_summary.items(), key=lambda x: -float(x[1]['total'])):
        payment_data.append([method, str(data['count']), f"${float(data['total']):,.0f}"])
    
    if len(payment_data) > 1:
        payment_table = Table(payment_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        payment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0EA5E9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), header_text),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(payment_table)
    
    # ========== CUENTAS POR COBRAR ==========
    if pending_contracts:
        elements.append(Paragraph("CUENTAS POR COBRAR (CONTRATOS VENCIDOS)", section_style))
        
        pending_data = [['Cliente', 'Vehículos/Combo', 'Vencimiento', 'Tarifa']]
        for c in pending_contracts:
            if c.use_combo_rate and c.combo_name:
                vehicles = f"[COMBO] {c.combo_name}"
            else:
                vehicles = c.vehicles_list[:25] + '..' if len(c.vehicles_list or '') > 25 else (c.vehicles_list or '-')
            pending_data.append([
                c.third_party.full_name[:20],
                vehicles,
                c.end_date.strftime('%d/%m/%Y'),
                f"${float(c.monthly_rate):,.0f}"
            ])
        pending_data.append(['', '', 'TOTAL', f"${total_pending:,.0f}"])
        
        pending_table = Table(pending_data, colWidths=[2*inch, 2*inch, 1*inch, 1*inch])
        pending_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), header_text),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (-1, -1), (-1, -1), red_text),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(pending_table)
    
    # ========== DETALLE DE TICKETS (últimos 30) ==========
    if tickets:
        elements.append(PageBreak())
        elements.append(Paragraph("DETALLE DE TICKETS DE PARQUEADERO", section_style))
        
        ticket_data = [['Fecha', 'Placa', 'Categoría', 'Método', 'Monto']]
        for t in tickets[:30]:
            ticket_data.append([
                t.exit_time.strftime('%d/%m %H:%M'),
                t.placa,
                t.category.name[:10] if t.category else '-',
                (t.payment_method.name[:10] if t.payment_method else 'Efectivo'),
                f"${float(t.amount_paid or 0):,.0f}"
            ])
        
        if tickets.count() > 30:
            ticket_data.append([f'... y {tickets.count() - 30} más', '', '', '', ''])
        
        ticket_data.append(['', '', '', 'TOTAL', f"${total_tickets:,.0f}"])
        
        ticket_table = Table(ticket_data, colWidths=[1.2*inch, 1*inch, 1.2*inch, 1.2*inch, 1*inch])
        ticket_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), header_text),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (-1, -1), (-1, -1), green_text),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(ticket_table)
    
    # ========== DETALLE DE GASTOS ==========
    if expenses:
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("DETALLE DE GASTOS", section_style))
        
        expense_detail_data = [['Fecha', 'Categoría', 'Descripción', 'Monto']]
        for e in expenses[:20]:
            expense_detail_data.append([
                e.date.strftime('%d/%m/%Y'),
                e.category.name[:15],
                (e.description[:25] + '..') if e.description and len(e.description) > 25 else (e.description or '-'),
                f"${float(e.amount or 0):,.0f}"
            ])
        
        if expenses.count() > 20:
            expense_detail_data.append([f'... y {expenses.count() - 20} más', '', '', ''])
        
        expense_detail_data.append(['', '', 'TOTAL', f"${total_expenses:,.0f}"])
        
        expense_detail_table = Table(expense_detail_data, colWidths=[1*inch, 1.5*inch, 2.5*inch, 1*inch])
        expense_detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), header_text),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (-1, -1), (-1, -1), red_text),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(expense_detail_table)
    
    # ========== DETALLE DE MENSUALIDADES ==========
    if contract_payments:
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("PAGOS DE MENSUALIDADES", section_style))
        
        contract_data = [['Fecha', 'Cliente', 'Período', 'Monto']]
        for p in contract_payments:
            contract_data.append([
                p.payment_date.strftime('%d/%m %H:%M'),
                str(p.contract.third_party)[:20] if p.contract else '-',
                f"{p.get_month_name()} {p.payment_year}",
                f"${float(p.amount or 0):,.0f}"
            ])
        contract_data.append(['', '', 'TOTAL', f"${total_contracts:,.0f}"])
        
        contract_table = Table(contract_data, colWidths=[1.2*inch, 2.5*inch, 1.3*inch, 1*inch])
        contract_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0EA5E9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), header_text),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (-1, -1), (-1, -1), green_text),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(contract_table)
    
    # Pie de página
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=1, textColor=colors.HexColor('#94A3B8'))
    elements.append(Paragraph(f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} - SoluPark", footer_style))
    
    doc.build(elements)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=reporte_{start_str}_a_{end_str}.pdf'
    response.write(buffer.getvalue())
    return response


def custom_logout(request):
    logout(request)
    messages.info(request, "Has cerrado sesión correctamente.")
    return redirect('login')


@login_required
def category_edit(request, pk):
    tenant = get_tenant_from_user(request.user)
    
    category_query = VehicleCategory.objects.all_tenants().filter(pk=pk)
    if tenant:
        category_query = category_query.filter(tenant=tenant)
    category = get_object_or_404(category_query)

    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoría actualizada correctamente.")
            return redirect('category-list')
        else:
            messages.error(request, "Hubo un error al actualizar la categoría.")
    else:
        form = CategoryForm(instance=category)

    return render(request, 'parking/category_edit.html', {'form': form})


@login_required
def validate_plate(request, plate):
    tenant = get_tenant_from_user(request.user)
    
    query = ParkingTicket.objects.all_tenants().filter(
        placa__iexact=plate,
        exit_time__isnull=True
    )
    if tenant:
        query = query.filter(tenant=tenant)
    
    exists = query.exists()
    return JsonResponse({'exists': exists})


class CategoryDeleteView(DeleteView):
    model = VehicleCategory
    success_url = reverse_lazy('category-list')
    template_name = 'parking/category_confirm_delete.html'

    def get_queryset(self):
        tenant = get_tenant_from_user(self.request.user)
        if tenant:
            return VehicleCategory.objects.all_tenants().filter(tenant=tenant)
        return VehicleCategory.objects.none()

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Categoría eliminada exitosamente')
        return super().delete(request, *args, **kwargs)


@login_required
def cash_register(request):
    tenant = get_tenant_from_user(request.user)
    is_vendedor = request.user.groups.filter(name='Vendedor').exists()
    today = timezone.now().date()
    
    # Obtener turno activo del usuario
    from .models import Turno
    turno_activo = Turno.objects.all_tenants().filter(
        tenant=tenant,
        user=request.user,
        is_active=True
    ).first() if tenant else None

    if is_vendedor:
        start_date = timezone.make_aware(datetime.combine(today, datetime.min.time()))
        end_date = start_date + timedelta(days=1)
    else:
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        if start_date_str and end_date_str:
            try:
                start_date = timezone.make_aware(datetime.strptime(f"{start_date_str} 00:00:00", '%Y-%m-%d %H:%M:%S'))
                end_date = timezone.make_aware(datetime.strptime(f"{end_date_str} 23:59:59", '%Y-%m-%d %H:%M:%S'))
            except ValueError:
                start_date = timezone.make_aware(datetime.combine(today, datetime.min.time()))
                end_date = start_date + timedelta(days=1)
        else:
            start_date = timezone.make_aware(datetime.combine(today, datetime.min.time()))
            end_date = start_date + timedelta(days=1)
    
    # Si hay turno activo, filtrar desde que se abrió el turno
    if turno_activo:
        start_date = turno_activo.start_time

    # Tickets de parqueadero
    if tenant:
        tickets = ParkingTicket.objects.all_tenants().filter(
            tenant=tenant,
            exit_time__gte=start_date,
            exit_time__lt=end_date,
            exit_time__isnull=False,
            amount_paid__isnull=False
        ).select_related('payment_method', 'category')
    else:
        tickets = ParkingTicket.objects.none()
    
    # Pagos de mensualidades
    from monthly_contracts.models import ContractPayment
    if tenant:
        contract_payments = ContractPayment.objects.filter(
            tenant=tenant,
            payment_date__gte=start_date,
            payment_date__lt=end_date,
            is_confirmed=True
        ).select_related('payment_method', 'contract', 'contract__third_party')
    else:
        contract_payments = ContractPayment.objects.none()
    
    # Calcular resumen por método de pago
    from parking.models_config import PaymentMethod
    from collections import defaultdict
    from decimal import Decimal
    
    payment_summary = defaultdict(lambda: {'count': 0, 'total': Decimal('0'), 'is_cash': False})
    
    # Sumar tickets de parqueadero
    for ticket in tickets:
        if ticket.payment_method:
            method_name = ticket.payment_method.name
            payment_summary[method_name]['count'] += 1
            payment_summary[method_name]['total'] += ticket.amount_paid or Decimal('0')
            payment_summary[method_name]['is_cash'] = ticket.payment_method.payment_type == 'cash'
            payment_summary[method_name]['icon'] = ticket.payment_method.icon
        else:
            # Sin método de pago asignado, asumir efectivo
            payment_summary['Efectivo (sin especificar)']['count'] += 1
            payment_summary['Efectivo (sin especificar)']['total'] += ticket.amount_paid or Decimal('0')
            payment_summary['Efectivo (sin especificar)']['is_cash'] = True
            payment_summary['Efectivo (sin especificar)']['icon'] = 'fa-money-bill'
    
    # Sumar pagos de mensualidades
    for payment in contract_payments:
        if payment.payment_method:
            method_name = payment.payment_method.name
            payment_summary[method_name]['count'] += 1
            payment_summary[method_name]['total'] += payment.amount or Decimal('0')
            payment_summary[method_name]['is_cash'] = payment.payment_method.payment_type == 'cash'
            payment_summary[method_name]['icon'] = payment.payment_method.icon
            payment_summary[method_name]['has_contracts'] = True
    
    # Convertir a lista ordenada
    payment_summary_list = [
        {'name': name, **data} 
        for name, data in payment_summary.items()
    ]
    payment_summary_list.sort(key=lambda x: (-x['total'], x['name']))
    
    # Calcular totales
    total_income = sum(ticket.amount_paid or 0 for ticket in tickets)
    total_contracts_cash = sum(
        p.amount for p in contract_payments 
        if p.payment_method and p.payment_method.payment_type == 'cash'
    )
    total_contracts = sum(p.amount for p in contract_payments)
    
    # Solo efectivo para caja física
    cash_from_tickets = sum(
        ticket.amount_paid or 0 for ticket in tickets 
        if ticket.payment_method and ticket.payment_method.payment_type == 'cash'
    ) + sum(
        ticket.amount_paid or 0 for ticket in tickets 
        if not ticket.payment_method  # Sin método = efectivo
    )
    
    total_cash = float(cash_from_tickets) + float(total_contracts_cash)
    total_all = float(total_income) + float(total_contracts)

    caja_date = start_date.date() if not turno_activo else today
    if tenant:
        caja_query = Caja.objects.all_tenants().filter(tenant=tenant, fecha=caja_date, tipo='Ingreso')
    else:
        caja_query = Caja.objects.none()

    if caja_query.exists():
        caja = caja_query.first()
        caja.monto = Decimal(str(total_cash))
        caja.save()
    else:
        caja = Caja(
            fecha=caja_date,
            tipo='Ingreso',
            monto=Decimal(str(total_cash)),
            descripcion=f'Ingresos del {caja_date}',
            dinero_inicial=0.00
        )
        if tenant:
            caja.tenant = tenant
        caja.save()

    # Usar dinero inicial del turno activo si existe
    dinero_inicial_base = float(turno_activo.initial_cash) if turno_activo else float(caja.dinero_inicial)
    
    # Obtener gastos del período
    from .models import Expense
    if tenant:
        expenses = Expense.objects.all_tenants().filter(
            tenant=tenant,
            date__gte=start_date.date(),
            date__lte=(end_date - timedelta(seconds=1)).date()
        ).select_related('category', 'created_by').order_by('-created_at')
    else:
        expenses = Expense.objects.none()
    
    total_expenses = sum(float(e.amount) for e in expenses)
    
    # Dinero esperado = base + ingresos efectivo - gastos
    dinero_esperado = dinero_inicial_base + total_cash - total_expenses

    if request.method == 'POST' and 'set_dinero_inicial' in request.POST:
        try:
            dinero_inicial = float(request.POST.get('dinero_inicial', 0))
            if dinero_inicial < 0:
                messages.error(request, 'El dinero inicial no puede ser negativo.')
            else:
                caja.dinero_inicial = dinero_inicial
                if turno_activo:
                    turno_activo.initial_cash = dinero_inicial
                    turno_activo.save()
                caja.save()
                messages.success(request, 'Dinero inicial establecido correctamente.')
            return redirect('cash_register')
        except ValueError:
            messages.error(request, 'Por favor, ingrese un valor numérico válido.')

    if request.method == 'POST' and 'realizar_cuadre' in request.POST:
        if caja.cuadre_realizado:
            messages.error(request, 'El cuadre de caja ya fue realizado.')
            return redirect('cash_register')

        try:
            dinero_final = float(request.POST.get('dinero_final', 0))
            if dinero_final < 0:
                messages.error(request, 'El dinero final no puede ser negativo.')
                return redirect('cash_register')

            caja.dinero_final = dinero_final
            caja.monto = Decimal(str(total_cash))
            caja.cuadre_realizado = True
            caja.save()
            
            # Cerrar el turno activo del usuario
            from .models import Turno
            turno_activo = Turno.objects.all_tenants().filter(
                tenant=tenant,
                user=request.user,
                is_active=True
            ).first()
            
            if turno_activo:
                turno_activo.end_time = timezone.now()
                turno_activo.final_cash = dinero_final
                turno_activo.expected_cash = dinero_esperado
                turno_activo.difference = dinero_final - dinero_esperado
                turno_activo.is_active = False
                turno_activo.closed_by = request.user
                turno_activo.save()

            messages.success(request, 'Cuadre de caja realizado con éxito. Turno cerrado.')
            return redirect('cash_register')
        except ValueError:
            messages.error(request, 'Por favor, ingrese un valor numérico válido.')

    diferencia = None
    diferencia_abs = None
    if caja.cuadre_realizado:
        diferencia = float(caja.dinero_final) - float(caja.dinero_inicial) - total_cash
        diferencia_abs = abs(diferencia)

    # Calcular promedio por ticket
    avg_ticket = total_income / len(tickets) if tickets else 0
    total_tickets = total_income

    context = {
        'today': today,
        'start_date': start_date.date(),
        'end_date': (end_date - timedelta(days=1)).date(),
        'tickets': tickets,
        'contract_payments': contract_payments,
        'payment_summary': payment_summary_list,
        'total_income': total_income,
        'total_tickets': total_tickets,
        'total_contracts': total_contracts,
        'total_cash': total_cash,
        'total_all': total_all,
        'avg_ticket': avg_ticket,
        'caja': caja,
        'dinero_esperado': dinero_esperado,
        'diferencia': diferencia,
        'diferencia_abs': diferencia_abs,
        'is_vendedor': is_vendedor,
        'tenant': tenant,
        'expenses': expenses,
        'total_expenses': total_expenses,
    }
    return render(request, 'parking/cash_register.html', context)


@login_required
def export_cash_register_excel(request):
    """Exportar cuadre de caja a Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from django.http import HttpResponse
    
    tenant = get_tenant_from_user(request.user)
    date_str = request.GET.get('date', timezone.now().date().isoformat())
    
    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        report_date = timezone.now().date()
    
    start_date = timezone.make_aware(datetime.combine(report_date, datetime.min.time()))
    end_date = start_date + timedelta(days=1)
    
    # Obtener datos
    tickets = ParkingTicket.objects.all_tenants().filter(
        tenant=tenant,
        exit_time__gte=start_date,
        exit_time__lt=end_date,
        exit_time__isnull=False,
        amount_paid__isnull=False
    ).select_related('payment_method', 'category')
    
    from monthly_contracts.models import ContractPayment
    contract_payments = ContractPayment.objects.filter(
        tenant=tenant,
        payment_date__gte=start_date,
        payment_date__lt=end_date,
        is_confirmed=True
    ).select_related('payment_method', 'contract', 'contract__third_party')
    
    # Obtener gastos
    from .models import Expense
    expenses = Expense.objects.all_tenants().filter(
        tenant=tenant,
        date=report_date
    ).select_related('category', 'created_by').order_by('-created_at')
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuadre de Caja"
    
    # Estilos
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    expense_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = f"CUADRE DE CAJA - {tenant.name if tenant else 'SoluPark'}"
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Fecha: {report_date.strftime('%d/%m/%Y')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Tickets de parqueadero
    row = 4
    ws[f'A{row}'] = "TICKETS DE PARQUEADERO"
    ws[f'A{row}'].font = subheader_font
    
    row += 1
    headers = ['Placa', 'Categoría', 'Entrada', 'Salida', 'Método Pago', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
    
    total_tickets = 0
    for ticket in tickets:
        row += 1
        ws.cell(row=row, column=1, value=ticket.placa).border = border
        ws.cell(row=row, column=2, value=ticket.category.name if ticket.category else '').border = border
        ws.cell(row=row, column=3, value=ticket.entry_time.strftime('%H:%M')).border = border
        ws.cell(row=row, column=4, value=ticket.exit_time.strftime('%H:%M') if ticket.exit_time else '').border = border
        ws.cell(row=row, column=5, value=ticket.payment_method.name if ticket.payment_method else 'Efectivo').border = border
        ws.cell(row=row, column=6, value=float(ticket.amount_paid or 0)).border = border
        total_tickets += float(ticket.amount_paid or 0)
    
    row += 1
    ws.cell(row=row, column=5, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=total_tickets).font = Font(bold=True)
    
    # Pagos de mensualidades
    row += 2
    ws[f'A{row}'] = "PAGOS DE MENSUALIDADES"
    ws[f'A{row}'].font = subheader_font
    
    row += 1
    headers = ['Cliente', 'Vehículos', 'Método Pago', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
    
    total_contracts = 0
    for payment in contract_payments:
        row += 1
        ws.cell(row=row, column=1, value=str(payment.contract.third_party) if payment.contract and payment.contract.third_party else '').border = border
        ws.cell(row=row, column=2, value=payment.contract.vehicles_list if payment.contract else '').border = border
        ws.cell(row=row, column=3, value=payment.payment_method.name if payment.payment_method else '').border = border
        ws.cell(row=row, column=4, value=float(payment.amount or 0)).border = border
        total_contracts += float(payment.amount or 0)
    
    row += 1
    ws.cell(row=row, column=3, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_contracts).font = Font(bold=True)
    
    # Gastos
    row += 2
    ws[f'A{row}'] = "GASTOS / EGRESOS"
    ws[f'A{row}'].font = subheader_font
    
    row += 1
    headers = ['Hora', 'Categoría', 'Descripción', 'Registrado por', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = expense_fill
        cell.border = border
    
    total_expenses = 0
    for expense in expenses:
        row += 1
        ws.cell(row=row, column=1, value=expense.created_at.strftime('%H:%M')).border = border
        ws.cell(row=row, column=2, value=expense.category.name).border = border
        ws.cell(row=row, column=3, value=expense.description).border = border
        ws.cell(row=row, column=4, value=expense.created_by.get_full_name() if expense.created_by else '').border = border
        ws.cell(row=row, column=5, value=float(expense.amount)).border = border
        total_expenses += float(expense.amount)
    
    row += 1
    ws.cell(row=row, column=4, value="TOTAL GASTOS:").font = Font(bold=True, color="DC2626")
    ws.cell(row=row, column=5, value=total_expenses).font = Font(bold=True, color="DC2626")
    
    # Resumen final
    row += 2
    ws[f'A{row}'] = "RESUMEN"
    ws[f'A{row}'].font = subheader_font
    
    row += 1
    ws.cell(row=row, column=1, value="Total Ingresos:").border = border
    ws.cell(row=row, column=2, value=total_tickets + total_contracts).border = border
    row += 1
    ws.cell(row=row, column=1, value="Total Gastos:").border = border
    ws.cell(row=row, column=2, value=total_expenses).font = Font(color="DC2626")
    row += 1
    ws.cell(row=row, column=1, value="NETO:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_tickets + total_contracts - total_expenses).font = Font(bold=True)
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    
    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=cuadre_caja_{report_date}.xlsx'
    wb.save(response)
    return response


@login_required
def export_cash_register_pdf(request):
    """Exportar cuadre de caja a PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from django.http import HttpResponse
    from io import BytesIO
    
    tenant = get_tenant_from_user(request.user)
    date_str = request.GET.get('date', timezone.now().date().isoformat())
    
    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        report_date = timezone.now().date()
    
    start_date = timezone.make_aware(datetime.combine(report_date, datetime.min.time()))
    end_date = start_date + timedelta(days=1)
    
    # Obtener datos
    tickets = ParkingTicket.objects.all_tenants().filter(
        tenant=tenant,
        exit_time__gte=start_date,
        exit_time__lt=end_date,
        exit_time__isnull=False,
        amount_paid__isnull=False
    ).select_related('payment_method', 'category')
    
    from monthly_contracts.models import ContractPayment
    contract_payments = ContractPayment.objects.filter(
        tenant=tenant,
        payment_date__gte=start_date,
        payment_date__lt=end_date,
        is_confirmed=True
    ).select_related('payment_method', 'contract', 'contract__third_party')
    
    # Obtener gastos
    from .models import Expense
    expenses = Expense.objects.all_tenants().filter(
        tenant=tenant,
        date=report_date
    ).select_related('category', 'created_by').order_by('-created_at')
    
    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1,
        spaceAfter=12
    )
    elements.append(Paragraph(f"CUADRE DE CAJA", title_style))
    elements.append(Paragraph(f"{tenant.name if tenant else 'SoluPark'}", styles['Heading2']))
    elements.append(Paragraph(f"Fecha: {report_date.strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Tabla de tickets
    elements.append(Paragraph("Tickets de Parqueadero", styles['Heading3']))
    
    ticket_data = [['Placa', 'Categoría', 'Entrada', 'Salida', 'Método', 'Monto']]
    total_tickets = 0
    for ticket in tickets:
        ticket_data.append([
            ticket.placa,
            ticket.category.name if ticket.category else '',
            ticket.entry_time.strftime('%H:%M'),
            ticket.exit_time.strftime('%H:%M') if ticket.exit_time else '',
            ticket.payment_method.name if ticket.payment_method else 'Efectivo',
            f"${float(ticket.amount_paid or 0):,.0f}"
        ])
        total_tickets += float(ticket.amount_paid or 0)
    
    ticket_data.append(['', '', '', '', 'TOTAL:', f"${total_tickets:,.0f}"])
    
    ticket_table = Table(ticket_data, colWidths=[1*inch, 1.2*inch, 0.8*inch, 0.8*inch, 1*inch, 1*inch])
    ticket_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0EA5E9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(ticket_table)
    elements.append(Spacer(1, 20))
    
    # Tabla de mensualidades
    elements.append(Paragraph("Pagos de Mensualidades", styles['Heading3']))
    
    contract_data = [['Cliente', 'Vehículos', 'Método', 'Monto']]
    total_contracts = 0
    for payment in contract_payments:
        contract_data.append([
            str(payment.contract.third_party)[:20] if payment.contract and payment.contract.third_party else '',
            (payment.contract.vehicles_list[:15] + '..') if payment.contract and len(payment.contract.vehicles_list) > 15 else (payment.contract.vehicles_list if payment.contract else ''),
            payment.payment_method.name if payment.payment_method else '',
            f"${float(payment.amount or 0):,.0f}"
        ])
        total_contracts += float(payment.amount or 0)
    
    contract_data.append(['', '', 'TOTAL:', f"${total_contracts:,.0f}"])
    
    contract_table = Table(contract_data, colWidths=[1.2*inch, 2*inch, 1.2*inch, 1*inch])
    contract_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(contract_table)
    elements.append(Spacer(1, 20))
    
    # Tabla de gastos
    elements.append(Paragraph("Gastos / Egresos", styles['Heading3']))
    
    expense_data = [['Hora', 'Categoría', 'Descripción', 'Monto']]
    total_expenses = 0
    for expense in expenses:
        expense_data.append([
            expense.created_at.strftime('%H:%M'),
            expense.category.name,
            expense.description[:25] + '..' if len(expense.description) > 25 else expense.description,
            f"-${float(expense.amount):,.0f}"
        ])
        total_expenses += float(expense.amount)
    
    expense_data.append(['', '', 'TOTAL:', f"-${total_expenses:,.0f}"])
    
    expense_table = Table(expense_data, colWidths=[0.8*inch, 1.5*inch, 2*inch, 1*inch])
    expense_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (-1, 1), (-1, -1), colors.HexColor('#DC2626')),
    ]))
    elements.append(expense_table)
    elements.append(Spacer(1, 20))
    
    # Resumen
    elements.append(Paragraph("RESUMEN", styles['Heading3']))
    summary_data = [
        ['Concepto', 'Monto'],
        ['Total Ingresos', f"${total_tickets + total_contracts:,.0f}"],
        ['Total Gastos', f"-${total_expenses:,.0f}"],
        ['NETO', f"${total_tickets + total_contracts - total_expenses:,.0f}"],
    ]
    summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(summary_table)
    
    doc.build(elements)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=cuadre_caja_{report_date}.pdf'
    response.write(buffer.getvalue())
    return response


# ==========================================
# SISTEMA DE TURNOS
# ==========================================

from .models import Turno


def get_active_turno(user, tenant):
    """Obtiene el turno activo del usuario"""
    if not tenant:
        return None
    return Turno.objects.all_tenants().filter(
        tenant=tenant,
        user=user,
        is_active=True
    ).first()


def has_active_turno(user, tenant):
    """Verifica si el usuario tiene un turno activo"""
    return get_active_turno(user, tenant) is not None


@login_required
def abrir_turno(request):
    """Vista para abrir un nuevo turno - pide dinero inicial y enlaza con caja"""
    tenant = get_tenant_from_user(request.user)
    
    if not tenant:
        messages.error(request, 'No tiene un parqueadero asignado.')
        return redirect('dashboard')
    
    # Verificar si ya tiene turno activo
    turno_activo = get_active_turno(request.user, tenant)
    if turno_activo:
        messages.warning(request, 'Ya tiene un turno activo.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        try:
            initial_cash = float(request.POST.get('initial_cash', 0))
            if initial_cash < 0:
                messages.error(request, 'El dinero inicial no puede ser negativo.')
                return redirect('abrir_turno')
            
            # Buscar caja del día
            today = timezone.now().date()
            caja = Caja.objects.all_tenants().filter(
                tenant=tenant,
                fecha=today,
                tipo='Ingreso'
            ).first()
            
            # Si la caja existe y ya tiene cuadre realizado, resetearla para el nuevo turno
            if caja and caja.cuadre_realizado:
                caja.cuadre_realizado = False
                caja.dinero_inicial = initial_cash
                caja.dinero_final = None
                caja.monto = 0
                caja.save()
            elif not caja:
                caja = Caja(
                    tenant=tenant,
                    fecha=today,
                    tipo='Ingreso',
                    dinero_inicial=initial_cash,
                    monto=0,
                    descripcion=f'Caja del {today}'
                )
                caja.save()
            elif not caja.dinero_inicial or caja.dinero_inicial == 0:
                caja.dinero_inicial = initial_cash
                caja.save()
            
            # Crear el turno enlazado a la caja
            turno = Turno(
                user=request.user,
                initial_cash=initial_cash,
                caja=caja,
                tenant=tenant
            )
            turno.save()
            
            messages.success(request, f'Turno abierto con ${initial_cash:,.0f} de base.')
            return redirect('dashboard')
        except ValueError:
            messages.error(request, 'Por favor ingrese un valor numérico válido.')
    
    return render(request, 'parking/abrir_turno.html', {
        'tenant': tenant
    })



# ========== VISTAS DE GASTOS ==========

@login_required
def expense_list(request):
    """Lista de gastos del día o período seleccionado"""
    tenant = get_tenant_from_user(request.user)
    today = timezone.now().date()
    
    # Filtros de fecha
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = end_date = today
    else:
        start_date = end_date = today
    
    from .models import Expense, ExpenseCategory
    
    if tenant:
        expenses = Expense.objects.all_tenants().filter(
            tenant=tenant,
            date__gte=start_date,
            date__lte=end_date
        ).select_related('category', 'created_by').order_by('-date', '-created_at')
        
        categories = ExpenseCategory.objects.all_tenants().filter(
            tenant=tenant,
            is_active=True
        ).order_by('name')
    else:
        expenses = Expense.objects.none()
        categories = ExpenseCategory.objects.none()
    
    # Calcular totales
    total_expenses = sum(e.amount for e in expenses)
    
    # Agrupar por categoría
    from collections import defaultdict
    by_category = defaultdict(lambda: {'count': 0, 'total': 0})
    for e in expenses:
        by_category[e.category.name]['count'] += 1
        by_category[e.category.name]['total'] += float(e.amount)
    
    context = {
        'expenses': expenses,
        'categories': categories,
        'start_date': start_date,
        'end_date': end_date,
        'total_expenses': total_expenses,
        'by_category': dict(by_category),
        'tenant': tenant,
    }
    return render(request, 'parking/expenses.html', context)


@login_required
def expense_create(request):
    """Crear un nuevo gasto"""
    tenant = get_tenant_from_user(request.user)
    
    from .models import Expense, ExpenseCategory, Turno
    
    # Verificar turno activo
    turno_activo = None
    if tenant:
        turno_activo = Turno.objects.all_tenants().filter(
            tenant=tenant,
            user=request.user,
            is_active=True
        ).first()
    
    if request.method == 'POST':
        category_id = request.POST.get('category')
        amount = request.POST.get('amount')
        description = request.POST.get('description', '').strip()
        notes = request.POST.get('notes', '').strip()
        
        if not category_id or not amount:
            messages.error(request, 'Categoría y monto son obligatorios.')
            return redirect('expense_list')
        
        try:
            category = ExpenseCategory.objects.all_tenants().get(
                pk=category_id,
                tenant=tenant,
                is_active=True
            )
            amount = float(amount)
            
            if amount <= 0:
                messages.error(request, 'El monto debe ser mayor a cero.')
                return redirect('expense_list')
            
            expense = Expense(
                category=category,
                amount=amount,
                description=description or category.name,
                notes=notes,
                created_by=request.user,
                turno=turno_activo,
                tenant=tenant
            )
            expense.save()
            
            messages.success(request, f'Gasto de ${amount:,.0f} registrado correctamente.')
            return redirect('expense_list')
            
        except ExpenseCategory.DoesNotExist:
            messages.error(request, 'Categoría no válida.')
        except ValueError:
            messages.error(request, 'Monto no válido.')
    
    return redirect('expense_list')


@login_required
def expense_delete(request, pk):
    """Eliminar un gasto"""
    tenant = get_tenant_from_user(request.user)
    
    from .models import Expense
    
    if request.method == 'POST':
        try:
            expense = Expense.objects.all_tenants().get(pk=pk, tenant=tenant)
            expense.delete()
            messages.success(request, 'Gasto eliminado correctamente.')
        except Expense.DoesNotExist:
            messages.error(request, 'Gasto no encontrado.')
    
    return redirect('expense_list')


@login_required
def expense_category_list(request):
    """Lista y gestión de categorías de gastos"""
    tenant = get_tenant_from_user(request.user)
    
    from .models import ExpenseCategory
    
    if tenant:
        categories = ExpenseCategory.objects.all_tenants().filter(
            tenant=tenant
        ).order_by('name')
    else:
        categories = ExpenseCategory.objects.none()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            
            if not name:
                messages.error(request, 'El nombre es obligatorio.')
            elif ExpenseCategory.objects.all_tenants().filter(tenant=tenant, name__iexact=name).exists():
                messages.error(request, 'Ya existe una categoría con ese nombre.')
            else:
                category = ExpenseCategory(
                    name=name,
                    description=description,
                    tenant=tenant
                )
                category.save()
                messages.success(request, f'Categoría "{name}" creada correctamente.')
            return redirect('expense_category_list')
        
        elif action == 'toggle':
            category_id = request.POST.get('category_id')
            try:
                category = ExpenseCategory.objects.all_tenants().get(pk=category_id, tenant=tenant)
                category.is_active = not category.is_active
                category.save()
                status = 'activada' if category.is_active else 'desactivada'
                messages.success(request, f'Categoría "{category.name}" {status}.')
            except ExpenseCategory.DoesNotExist:
                messages.error(request, 'Categoría no encontrada.')
            return redirect('expense_category_list')
        
        elif action == 'delete':
            category_id = request.POST.get('category_id')
            try:
                category = ExpenseCategory.objects.all_tenants().get(pk=category_id, tenant=tenant)
                if category.expenses.exists():
                    messages.error(request, 'No se puede eliminar una categoría con gastos asociados.')
                else:
                    category.delete()
                    messages.success(request, 'Categoría eliminada correctamente.')
            except ExpenseCategory.DoesNotExist:
                messages.error(request, 'Categoría no encontrada.')
            return redirect('expense_category_list')
    
    context = {
        'categories': categories,
        'tenant': tenant,
    }
    return render(request, 'parking/expense_categories.html', context)
